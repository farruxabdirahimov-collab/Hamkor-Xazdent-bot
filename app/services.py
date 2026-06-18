import asyncio
import re
import os
import logging
import json as _json
from datetime import datetime, timedelta
from aiogram import F
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, WebAppInfo,
    BufferedInputFile,
)
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from app.runtime import bot, dp, router, buyer_bot
from app.config import (BOT_TOKEN, CHANNEL_ID, ADMIN_IDS, WEBAPP_URL, BASE_DIR, AD_PRICE_TOSHKENT, AD_PRICE_REGION, AD_PRICE_BOTH_AUD, AD_REGION_PRICES, AD_REGION_DEFAULT)
from app.database import (init_db, get_user, db_run, db_get, db_all, db_insert, get_setting, update_setting, add_balance, get_next_room_code, generate_order_number, log_order_event, get_or_create_trust_score, update_trust_score, update_seller_metrics, get_pool)
from app.texts import t, REGIONS, REGIONS_RU
from app.keyboards import (ib, ik, kb_cancel, kb_clinic, kb_confirm, kb_deadline, kb_delivery, kb_lang, kb_regions, kb_role, kb_seller, kb_shop_cats, kb_units, rk)
from app.states import (AdState, AdminState, BulkState, CheckoutState, ComplaintState, MyProductsState, NeedState, OfferState, PhotoOrderState, QuickOrderState, RegState, ReviewState, ShopState, SupportState, TopupState)
log = logging.getLogger(__name__)


async def lang(uid):
    u = await get_user(uid)
    return (u["lang"] if u else None) or "uz"

async def has_profile(uid):
    u = await get_user(uid)
    return bool(u and u["clinic_name"] and u["phone"] and u["region"])

async def get_or_create_room(uid):
    """Foydalanuvchining default omborxonasini topadi yoki yaratadi (ko'rinmaydi)."""
    room = await db_get(
        "SELECT * FROM rooms WHERE owner_id=? AND status='active' ORDER BY id LIMIT 1", (uid,)
    )
    if room:
        return room
    rid = await db_insert(
        "INSERT INTO rooms(room_code,room_type,owner_id,max_needs) VALUES(?,?,?,?)",
        (f"AUTO{uid}", "premium", uid, 9999),
    )
    return await db_get("SELECT * FROM rooms WHERE id=?", (rid,))

async def _handle_web_cart(msg: Message, state: FSMContext, u, cart_data: str):
    """Web saytdan savat buyurtmasi. Format: 123x2,456x1"""
    uid = msg.from_user.id

    # Savat ma'lumotlarini parse qilish
    items_raw = []
    try:
        for part in cart_data.split(','):
            if 'x' in part:
                pid, qty = part.split('x', 1)
                items_raw.append({'pid': int(pid), 'qty': int(qty)})
    except Exception:
        items_raw = []

    if not items_raw:
        await msg.answer("⚠️ Savat bo'sh yoki xato. Qayta urinib ko'ring.")
        return

    # Ro'yxatdan o'tmagan bo'lsa
    if not u or not u.get('phone') or not u.get('region'):
        # Savatni state ga saqlaymiz
        await state.update_data(web_cart=cart_data)
        if not u:
            await db_run(
                "INSERT INTO users(id,username,full_name) VALUES(?,?,?) ON CONFLICT(id) DO NOTHING",
                (uid, msg.from_user.username, msg.from_user.full_name)
            )
        lg = 'uz'
        await state.set_state(RegState.phone)
        kb = rk([KeyboardButton(text="📱 Telefon yuborish", request_contact=True)], one_time=True)
        await msg.answer(
            "🛒 *Savatdagi buyurtmangiz tayyor!*\n\n"
            "Buyurtmani rasmiylashtirish uchun\n"
            "telefon raqamingizni yuboring:",
            reply_markup=kb
        )
        return

    # Ro'yxatdan o'tgan — buyurtmani yaratamiz
    await _create_web_cart_order(msg, u, items_raw)

async def _create_web_cart_order(msg, u, items_raw):
    """Web cart dan buyurtma yaratish."""
    uid = msg.from_user.id
    uname = u.get("clinic_name") or u.get("full_name") or str(uid)
    uphone = u.get("phone") or "—"
    uregion = u.get("region") or "—"
    import json as _pj2
    seller_map = {}
    for item in items_raw:
        prod = await db_get(
            "SELECT p.*, s.owner_id as seller_id, s.shop_name "
            "FROM products p JOIN shops s ON p.shop_id=s.id WHERE p.id=?",
            (item["pid"],)
        )
        if not prod: continue
        sid = prod["seller_id"]
        if sid not in seller_map:
            seller_map[sid] = {"shop_name": prod["shop_name"], "items": []}
        seller_map[sid]["items"].append({
            "name": prod["name"], "qty": item["qty"],
            "price": prod["price"], "unit": prod["unit"],
            "product_id": prod["id"]
        })
    if not seller_map:
        await msg.answer("Mahsulotlar topilmadi. Katalog yangilangan bolishi mumkin.")
        return
    sent = 0
    for seller_id, data in seller_map.items():
        items = data["items"]
        total = sum(i["price"] * i["qty"] for i in items)
        lines_list = []
        for i, it in enumerate(items):
            lines_list.append(
                f"{i+1}. *{it['name']}* — {it['qty']} {it['unit']} x "
                f"{fmt_price(it['price'])} = *{fmt_price(it['price']*it['qty'])} som*"
            )
        lines_txt = "\n".join(lines_list)
        order_id = await db_insert(
            "INSERT INTO catalog_orders(buyer_id,seller_id,products_json,total_amount) VALUES(?,?,?,?)",
            (uid, seller_id, _pj2.dumps(items, ensure_ascii=False), total)
        )
        msg_txt = (
            f"🌐 *Veb-saytdan buyurtma #{order_id}!*\n\n"
            f"📦 *{data['shop_name']}:*\n{lines_txt}\n\n"
            f"💰 *Jami: {fmt_price(total)} som*\n\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🏥 *{uname}*\n"
            f"📞 {uphone}\n"
            f"📍 {uregion}"
        )
        confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ Qabul qildim", callback_data=f"co_confirm_{order_id}_{uid}"),
            InlineKeyboardButton(text="❌ Mavjud emas", callback_data=f"co_reject_{order_id}_{uid}")
        ]])
        try:
            await bot.send_message(seller_id, msg_txt, reply_markup=confirm_kb)
            sent += 1
        except Exception as e:
            log.error(f"Web cart seller notify: {e}")
    lg = await lang(uid)
    if not u.get("role") or u["role"] in (None,"none",""):
        await db_run("UPDATE users SET role=? WHERE id=?", ("clinic", uid))
        u2 = await get_user(uid)
    else:
        u2 = u
    kb = kb_clinic(lg) if u2 and u2.get("role") in ("clinic","zubtex") else kb_seller(lg)
    await msg.answer(
        f"✅ *Buyurtmangiz yuborildi!*\n\n"
        f"📦 {sent} ta sotuvchiga xabar ketdi.\n"
        f"Sotuvchilar tez orada boglanadi.",
        reply_markup=kb
    )


def fmt_price(n):
    return f"{int(n):,}".replace(',', ' ')

async def post_to_channel(need_id, need):
    """1 ta ehtiyoj uchun kanal posti (qayta post uchun)."""
    dl_map = {2:"2 soat",24:"24 soat",72:"3 kun",168:"1 hafta",240:"10 kun"}
    dl_txt = dl_map.get(need["deadline_hours"], f"{need['deadline_hours']} soat")
    owner  = await get_user(need["owner_id"])
    words  = need["product_name"].split()
    tags   = " ".join(f"#{w.lower()}" for w in words[:3] if len(w)>2)
    txt = (
        f"📋 *BUYURTMA #{need_id}*\n\n"
        f"🦷 {need['product_name']}\n"
        f"📦 {need['quantity']} {need['unit']}\n"
        f"⏱ {dl_txt} ichida\n\n"
        f"📍 {owner['region'] or ''}\n\n"
        f"{tags}\n💬 @XazdentBot"
    )
    batch_id = need.get("batch_id")
    # Kanalga WebAppInfo ishlamaydi — deep link ishlatamiz
    bot_info = await bot.get_me()
    deep_url = f"https://t.me/{bot_info.username}?start=offer_{need_id}"
    kb = ik([ib("📤 Taklif yuborish", url=deep_url)])
    try:
        m = await bot.send_message(CHANNEL_ID, txt, reply_markup=kb)
        return m.message_id
    except Exception as e:
        log.error(f"❌ Kanal xato: {e}")
        return None

async def post_batch_to_channel(batch_id, needs_list, owner, photo_file_id=None):
    """Ko'p ehtiyoj uchun BITTA paket post."""
    if not needs_list:
        return None
    dl_map = {2:"2 soat",24:"24 soat",72:"3 kun",168:"1 hafta",240:"10 kun"}
    dl_txt = dl_map.get(needs_list[0]["deadline_hours"], "?")
    lines  = "\n".join([
        f"• {n['product_name']} — {n['quantity']} {n['unit']}"
        for n in needs_list[:15]
    ])
    if len(needs_list) > 15:
        lines += f"\n• ...va yana {len(needs_list)-15} ta"
    all_words = " ".join(n["product_name"] for n in needs_list[:5]).split()
    tags = " ".join(f"#{w.lower()}" for w in dict.fromkeys(all_words) if len(w)>2)[:80]
    # To'lov turlari
    pay_icons = {"p2p":"💳 P2P","cash":"💵 Naqd","bank":"🏦 Hisob raqam"}
    pm_raw = needs_list[0].get("payment_methods","") if needs_list else ""
    pm_txt = " · ".join(pay_icons[p] for p in (pm_raw or "").split(",") if p in pay_icons)
    pm_line = f"\n💳 {pm_txt}" if pm_txt else ""
    txt = (
        f"📋 *BUYURTMA #{batch_id}* — {len(needs_list)} ta mahsulot\n\n"
        f"{lines}\n\n"
        f"📍 {owner.get('region') or ''}\n"
        f"⏱ {dl_txt} ichida{pm_line}\n\n"
        f"{tags}\n💬 @XazdentBot"
    )
    # Kanalga WebAppInfo yuborib bo'lmaydi — oddiy URL link ishlatamiz
    bot_info = await bot.get_me()
    if WEBAPP_URL:
        deep_url = f"https://t.me/{bot_info.username}?start=batch_{batch_id}"
    else:
        deep_url = f"https://t.me/{bot_info.username}?start=batch_{batch_id}"
    kb = ik([ib("📤 Taklif yuborish", url=deep_url)])
    try:
        if photo_file_id:
            m = await bot.send_photo(CHANNEL_ID, photo_file_id,
                                     caption=txt, reply_markup=kb)
        else:
            m = await bot.send_message(CHANNEL_ID, txt, reply_markup=kb)
        log.info(f"✅ Batch kanal post: batch={batch_id} msg={m.message_id}")
        return m.message_id
    except Exception as e:
        log.error(f"❌ Batch kanal xato: {e}")
        return None

async def notify_sellers(need_id, need, owner):
    """Barcha sotuvchilarga lichkada xabar."""
    sellers = await db_all(
        "SELECT id FROM users WHERE role='seller' AND id!=? AND is_blocked=0",
        (owner["id"],),
    )
    if WEBAPP_URL and need.get("batch_id"):
        url = f"{WEBAPP_URL}/offer/{need['batch_id']}"
        kb = ik(
            [ib("💰 Narx kiriting →", web_app=WebAppInfo(url=url))],
            [ib("⏭ Keyinroq", "skip_notify")],
        )
    else:
        bot_info = await bot.get_me()
        deep_url = f"https://t.me/{bot_info.username}?start=offer_{need_id}"
        kb = ik(
            [ib("📤 Taklif yuborish", url=deep_url)],
            [ib("⏭ Keyinroq", "skip_notify")],
        )

    dl_map = {2: "2 soat", 24: "24 soat", 72: "3 kun", 168: "1 hafta"}
    txt = (
        f"📦 *Yangi buyurtma!*\n\n"
        f"🦷 {need['product_name']}\n"
        f"📦 {need['quantity']} {need['unit']}\n"
        f"⏱ {dl_map.get(need['deadline_hours'], '?')} ichida\n"
        f"📍 {owner['region'] or ''}"
    )
    sent = 0
    for s in sellers:
        try:
            await bot.send_message(s["id"], txt, reply_markup=kb)
            sent += 1
            await asyncio.sleep(0.05)
        except Exception:
            pass
    log.info(f"Notify: {sent}/{len(sellers)} sotuvchiga yuborildi")

async def _finish_reg(msg: Message, state: FSMContext):
    uid = msg.from_user.id
    await state.clear()
    lg = await lang(uid)
    u  = await get_user(uid)
    if u and u["role"] in ("clinic", "zubtex"):
        kb    = kb_clinic(lg, uid=uid, webapp_url=WEBAPP_URL)
        txt   = (
            f"✅ *Ro\'yxatdan o\'tdingiz!*\n\n"
            f"🏥 {u.get('clinic_name','')}\n"
            f"📍 {u.get('region','')}\n\n"
            f"🛍 Dental Market orqali materiallar buyurtma qiling!"
        )
        await msg.answer(txt, reply_markup=kb)
    else:
        # Do'kon yo'q bo'lsa avtomatik yaratamiz
        shop = await db_get("SELECT id FROM shops WHERE owner_id=?", (uid,))
        if not shop and u:
            sname = u.get("clinic_name") or u.get("full_name") or "Do\'konim"
            await db_insert(
                "INSERT INTO shops(owner_id,shop_name,category,phone,region,status) "
                "VALUES(?,?,?,?,?,'active')",
                (uid, sname, "Stomatologiya", u.get("phone",""), u.get("region",""))
            )
        kb  = kb_seller(lg, uid=uid, webapp_url=WEBAPP_URL)
        txt = (
            f"✅ *Do\'koningiz ochildi!*\n\n"
            f"🏪 {u.get('clinic_name','')}\n"
            f"📍 {u.get('region','')}\n\n"
            f"➕ Mahsulot qo\'shing va xaridorlar sizni topsın!"
        )
        await msg.answer(txt, reply_markup=kb)
        # Adminga xabar
        for aid in ADMIN_IDS:
            try:
                await bot.send_message(
                    aid,
                    f"🆕 *Yangi sotuvchi!*\n\n"
                    f"🏪 {u.get('clinic_name','')}\n"
                    f"👤 {u.get('full_name','')}\n"
                    f"📞 {u.get('phone','')}\n"
                    f"📍 {u.get('region','')}\n"
                    f"🆔 ID: `{uid}`"
                )
            except Exception:
                pass

def _payment_kb(selected: list) -> InlineKeyboardMarkup:
    opts = [("p2p","💳 P2P (karta)"), ("bank","🏦 Hisob raqam"), ("cash","💵 Naqd pul")]
    rows = []
    for key, label in opts:
        chk = "✅ " if key in selected else "☐ "
        rows.append([ib(f"{chk}{label}", f"pm_tog_{key}")])
    rows.append([ib("💾 Saqlash", "pm_save"), ib("◀️ Bekor", "pm_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def notify_sellers_batch(batch_id: int, owner_id: int):
    """Batch dagi barcha ehtiyojlar haqida sotuvchilarga xabar."""
    needs   = await db_all("SELECT * FROM needs WHERE batch_id=?", (batch_id,))
    owner   = await get_user(owner_id)
    sellers = await db_all(
        "SELECT id FROM users WHERE role='seller' AND id!=? AND is_blocked=0", (owner_id,)
    )
    if not needs or not sellers:
        return

    preview = "\n".join([f"• {n['quantity']} {n['unit']} — {n['product_name']}" for n in needs[:5]])
    if len(needs) > 5:
        preview += f"\n• ...va yana {len(needs)-5} ta"

    if WEBAPP_URL:
        url = f"{WEBAPP_URL}/offer/{batch_id}"
        kb  = ik(
            [ib("💰 Narx kiriting →", web_app=WebAppInfo(url=url))],
            [ib("⏭ Keyinroq", "skip_notify")],
        )
    else:
        bot_info = await bot.get_me()
        kb = ik(
            [ib("📤 Taklif yuborish", url=f"https://t.me/{bot_info.username}?start=offer_{needs[0]['id']}")],
            [ib("⏭ Keyinroq", "skip_notify")],
        )

    dl_map = {2: "2 soat", 24: "24 soat", 72: "3 kun", 168: "1 hafta"}
    txt = (
        f"📦 *{len(needs)} ta buyurtma!*\n\n"
        f"{preview}\n\n"
        f"📍 {owner['region'] or ''}\n"
        f"⏱ {dl_map.get(needs[0]['deadline_hours'], '?')} ichida"
    )
    # Batch dagi rasmni topamiz
    photo_id = None
    for n in needs:
        if n.get("photo_file_id"):
            photo_id = n["photo_file_id"]
            break

    sent = 0
    for s in sellers:
        try:
            if photo_id:
                await bot.send_photo(s["id"], photo_id, caption=txt, reply_markup=kb)
            else:
                await bot.send_message(s["id"], txt, reply_markup=kb)
            sent += 1
            await asyncio.sleep(0.05)
        except Exception:
            pass
    log.info(f"Batch notify: {sent}/{len(sellers)} sotuvchiga")

async def _show_batch_table(target_msg, batch_id: int):
    """Batch dagi barcha takliflarni jadval ko'rinishida ko'rsatadi."""
    needs = await db_all(
        "SELECT * FROM needs WHERE batch_id=? AND status != 'cancelled' ORDER BY id",
        (batch_id,),
    )
    if not needs:
        await target_msg.answer("📭 Bu buyurtmada mahsulot yo'q.")
        return

    # Barcha sotuvchilarni topamiz
    sellers_map = {}  # seller_id → name
    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? AND o.price > 0 ORDER BY o.price ASC",
            (n["id"],),
        )
        for o in offs:
            sid = o["seller_id"]
            if sid not in sellers_map:
                sellers_map[sid] = o["clinic_name"] or o["full_name"] or f"Sotuvchi{sid}"

    has_offers = len(sellers_map) > 0
    txt = f"📊 *Jadval #{batch_id}*\n_{len(needs)} ta mahsulot"
    txt += f", {len(sellers_map)} ta taklif_\n\n" if has_offers else " — taklif kutilmoqda_\n\n"

    rows_for_accept = []  # (need_id, best_offer_id, best_price, seller_name)

    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? AND o.price > 0 ORDER BY o.price ASC",
            (n["id"],),
        )
        st = {"active":"🟢","paused":"⏸","done":"✅","cancelled":"❌"}.get(n["status"],"📋")
        txt += f"{st} *{n['product_name']}* — {n['quantity']} {n['unit']}\n"
        if offs:
            for i, o in enumerate(offs, 1):
                sname  = o["clinic_name"] or o["full_name"] or "Sotuvchi"
                marker = "✅ " if i == 1 else "   "
                note   = f" _{o['note']}_" if o.get("note") and o["note"] != "mavjud_emas" else ""
                txt   += f"  {marker}{i}. {sname} — {o['price']:,.0f} so'm{note}\n"
            if n["status"] == "active":
                best = offs[0]
                rows_for_accept.append((n["id"], best["id"], best["price"],
                                        best["clinic_name"] or best["full_name"] or "Sotuvchi"))
        else:
            txt += "  _taklif kelmagan_\n"
        txt += "\n"

    # Qabul tugmalari
    kb_rows = []
    if rows_for_accept:
        txt += "─────────────────\n"
        txt += "_Eng arzon taklif qabul qilish:_\n"
        for nid, oid, price, sname in rows_for_accept[:8]:
            short = sname[:12] + ("…" if len(sname) > 12 else "")
            kb_rows.append([ib(f"✅ {short} {price:,.0f}", f"acc_{oid}")])

    # Excel tugmasi
    kb_rows.append([ib("📥 Excel yuklab olish", f"xlsx_{batch_id}")])
    await target_msg.answer(txt, reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))

async def build_table(batch_id: int) -> str:
    """Batch dagi ehtiyojlar va takliflar jadvalini matn sifatida qaytaradi."""
    needs = await db_all(
        "SELECT * FROM needs WHERE batch_id=? ORDER BY id", (batch_id,)
    )
    if not needs:
        return "Bo'sh jadval."

    lines = [f"📊 *Jadval #{batch_id}*\n"]
    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? ORDER BY o.price ASC",
            (n["id"],),
        )
        lines.append(f"🦷 *{n['product_name']}* — {n['quantity']} {n['unit']}")
        if not offs:
            lines.append("   _Taklif kelmagan_")
        else:
            for i, o in enumerate(offs, 1):
                name   = o["clinic_name"] or o["full_name"] or "Sotuvchi"
                marker = "✅ " if i == 1 else ""
                lines.append(f"   {marker}{i}. {name} — {o['price']:,.0f} so'm")
        lines.append("")
    return "\n".join(lines)

async def build_excel(batch_id: int):
    """Excel fayl yaratadi, path qaytaradi."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    needs = await db_all("SELECT * FROM needs WHERE batch_id=?", (batch_id,))
    sellers_set = set()
    need_offers = {}
    for n in needs:
        offs = await db_all(
            "SELECT o.*, u.clinic_name, u.full_name FROM offers o "
            "JOIN users u ON o.seller_id=u.id "
            "WHERE o.need_id=? ORDER BY o.price ASC",
            (n["id"],),
        )
        need_offers[n["id"]] = list(offs)
        for o in offs:
            sellers_set.add(o["clinic_name"] or o["full_name"] or f"Sotuvchi{o['seller_id']}")

    sellers = sorted(sellers_set)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadval #{batch_id}"

    # Header
    header = ["Mahsulot", "Miqdor", "Birlik"] + [f"{s}\n(1 ta narx)" for s in sellers] + ["Eng arzon jami"]
    for col, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row_i, n in enumerate(needs, 2):
        ws.cell(row=row_i, column=1, value=n["product_name"])
        ws.cell(row=row_i, column=2, value=n["quantity"])
        ws.cell(row=row_i, column=3, value=n["unit"])
        offs      = need_offers[n["id"]]
        qty       = float(n["quantity"])
        min_total = None
        for o in offs:
            seller_name = o["clinic_name"] or o["full_name"] or f"Sotuvchi{o['seller_id']}"
            if seller_name in sellers:
                col_i   = sellers.index(seller_name) + 4
                unit_p  = o["price"]          # 1 ta uchun
                total_p = unit_p * qty        # jami
                # Katakda: "1 ta: 45,000" yozamiz, ustun sarlavhasida miqdor ko'rinadi
                ws.cell(row=row_i, column=col_i, value=unit_p)
                if min_total is None or total_p < min_total:
                    min_total = total_p
        if min_total is not None:
            last_col = len(sellers) + 4
            cell = ws.cell(row=row_i, column=last_col, value=min_total)
            cell.fill = PatternFill("solid", fgColor="E2EFDA")
            cell.font = Font(bold=True)

    # Column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    path = os.path.join(BASE_DIR, f"jadval_{batch_id}.xlsx")
    wb.save(path)
    return path

async def _start_offer_bot(msg_or_call, state: FSMContext, nid: int):
    uid = msg_or_call.from_user.id
    nd  = await db_get("SELECT * FROM needs WHERE id=?", (nid,))
    if not nd or nd["status"] != "active":
        txt = "⚠️ Bu ehtiyoj yopilgan yoki topilmadi."
        if hasattr(msg_or_call, "answer"):
            await msg_or_call.answer(txt)
        else:
            await msg_or_call.message.answer(txt)
        return

    exists = await db_get("SELECT id FROM offers WHERE need_id=? AND seller_id=?", (nid, uid))
    if exists:
        txt = "⚠️ Bu ehtiyojga allaqachon taklif yubordingiz!"
        if hasattr(msg_or_call, "answer"):
            await msg_or_call.answer(txt)
        else:
            await msg_or_call.message.answer(txt)
        return

    await state.update_data(need_id=nid, need_unit=nd["unit"], need_name=nd["product_name"])
    await state.set_state(OfferState.price)
    txt = (
        f"📦 *{nd['product_name']}* — {nd['quantity']} {nd['unit']}\n\n"
        f"💰 Narxingiz? _(1 {nd['unit']} uchun, so'mda)_"
    )
    no_stock = ik([ib("❌ Mavjud emas", f"no_stock_{nid}")])
    if hasattr(msg_or_call, "answer"):
        await msg_or_call.answer(txt, reply_markup=no_stock)
    else:
        await msg_or_call.message.answer(txt, reply_markup=no_stock)

async def _save_offer(obj, state: FSMContext, note):
    uid = obj.from_user.id
    d   = await state.get_data()
    u   = await get_user(uid)

    # Delivery default 24 soat
    await db_insert(
        "INSERT INTO offers(need_id,seller_id,product_name,price,unit,delivery_hours,note) VALUES(?,?,?,?,?,?,?)",
        (d["need_id"], uid, d["need_name"], d["price"], d["need_unit"], 24, note),
    )

    nd     = await db_get(
        "SELECT n.*, u2.id as cid FROM needs n JOIN users u2 ON n.owner_id=u2.id WHERE n.id=?",
        (d["need_id"],),
    )
    shop   = await db_get("SELECT shop_name FROM shops WHERE owner_id=? AND status='active'", (uid,))
    sname  = (shop["shop_name"] if shop else None) or u["clinic_name"] or u["full_name"] or "Sotuvchi"
    note_t = f"\n📝 _{note}_" if note else ""

    try:
        # Xaridor (klinika) XARIDOR botida — xabar buyer_bot orqali ketadi
        await buyer_bot.send_message(
            nd["cid"],
            f"📩 *Yangi taklif!*\n\n"
            f"🦷 {d['need_name']}\n"
            f"💰 *{d['price']:,.0f} so'm*/{d['need_unit']}\n"
            f"🏪 {sname}{note_t}",
            reply_markup=ik([ib(f"📩 Barcha takliflarni ko'rish", f"view_offers_{d['need_id']}")]),
        )
    except Exception as e:
        log.error(f"Klinikaga xabar xato: {e}")

    await state.clear()
    txt = f"✅ *Taklif yuborildi!*\n\n🦷 {d['need_name']}\n💰 {d['price']:,.0f} so'm/{d['need_unit']}"
    if note:
        txt += f"\n📝 {note}"
    if hasattr(obj, "answer"):
        await obj.answer(txt)
    else:
        await obj.message.answer(txt)

async def _show_seller_stats(uid: int, target_msg):
    now   = datetime.now()
    today = now.strftime("%Y-%m-%d")
    week  = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    month = now.strftime("%Y-%m")
    year  = now.strftime("%Y")

    async def won_sum(since=None, period=None):
        if period:
            rows = await db_all(
                "SELECT o.price, n.quantity FROM offers o "
                "JOIN needs n ON o.need_id=n.id "
                "WHERE o.seller_id=? AND o.status=\'accepted\' AND o.created_at LIKE ?",
                (uid, f"{period}%")
            )
        elif since:
            rows = await db_all(
                "SELECT o.price, n.quantity FROM offers o "
                "JOIN needs n ON o.need_id=n.id "
                "WHERE o.seller_id=? AND o.status=\'accepted\' AND o.created_at >= ?",
                (uid, since)
            )
        else:
            rows = await db_all(
                "SELECT o.price, n.quantity FROM offers o "
                "JOIN needs n ON o.need_id=n.id "
                "WHERE o.seller_id=? AND o.status=\'accepted\'",
                (uid,)
            )
        return sum(r["price"] * r["quantity"] for r in rows), len(rows)

    day_sum,   day_cnt   = await won_sum(since=today)
    week_sum,  week_cnt  = await won_sum(since=week)
    month_sum, month_cnt = await won_sum(period=month)
    year_sum,  year_cnt  = await won_sum(period=year)
    total_sum, total_cnt = await won_sum()

    # Jami takliflar
    all_offs  = await db_get("SELECT COUNT(*) as c FROM offers WHERE seller_id=?", (uid,))
    all_count = all_offs["c"] if all_offs else 0
    rate      = f"{total_cnt/all_count*100:.0f}%" if all_count else "—"

    # Top 5 mahsulot
    top_prods = await db_all(
        "SELECT n.product_name, COUNT(*) as cnt, SUM(o.price*n.quantity) as total "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "GROUP BY n.product_name ORDER BY total DESC LIMIT 5",
        (uid,)
    )

    txt = (
        f"📊 *Savdo statistikasi*\n\n"
        f"📅 Bugun:      *{day_sum:>12,.0f} so'm* ({day_cnt} ta)\n"
        f"📅 Bu hafta:   *{week_sum:>12,.0f} so'm* ({week_cnt} ta)\n"
        f"📅 Bu oy:      *{month_sum:>12,.0f} so'm* ({month_cnt} ta)\n"
        f"📅 Bu yil:     *{year_sum:>12,.0f} so'm* ({year_cnt} ta)\n"
        f"📅 Jami:       *{total_sum:>12,.0f} so'm* ({total_cnt} ta)\n\n"
        f"📤 Jami taklif: {all_count} ta | ✅ Qabul: {rate}\n"
    )

    if top_prods:
        txt += "\n🏆 *Top mahsulotlar:*\n"
        for i, p in enumerate(top_prods, 1):
            txt += f"  {i}. {p['product_name']} — {p['total']:,.0f} so'm ({p['cnt']} ta)\n"

    await target_msg.answer(txt, reply_markup=ik(
        [ib("📥 Excel yuklab olish", "seller_excel")],
        [ib("◀️ Orqaga", "seller_stats_back")],
    ))

async def _build_seller_excel(uid: int) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb  = openpyxl.Workbook()
    now = datetime.now()

    def hdr(ws, cols, color="4472C4"):
        fill = PatternFill("solid", fgColor=color)
        font = Font(bold=True, color="FFFFFF")
        for i, v in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=v)
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="center")

    def aw(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w+4, 40)

    # ── 1. Umumiy savdo ──────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Jami savdo"
    hdr(ws1, ["Sana", "Mahsulot", "Miqdor", "Birlik", "Narx (1 ta)", "Jami", "Klinika"])
    offs = await db_all(
        "SELECT o.*, n.product_name, n.quantity, n.unit, "
        "COALESCE(u.clinic_name, u.full_name) as clinic "
        "FROM offers o "
        "JOIN needs n ON o.need_id=n.id "
        "JOIN users u ON n.owner_id=u.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "ORDER BY o.created_at DESC",
        (uid,)
    )
    for i, o in enumerate(offs, 2):
        ws1.cell(row=i, column=1, value=o["created_at"][:10] if o["created_at"] else "")
        ws1.cell(row=i, column=2, value=o["product_name"])
        ws1.cell(row=i, column=3, value=o["quantity"])
        ws1.cell(row=i, column=4, value=o["unit"])
        ws1.cell(row=i, column=5, value=o["price"])
        ws1.cell(row=i, column=6, value=o["price"] * o["quantity"])
        ws1.cell(row=i, column=7, value=o["clinic"] or "—")
    # Jami
    if offs:
        row = len(offs) + 2
        ws1.cell(row=row, column=5, value="JAMI:").font = Font(bold=True)
        total = sum(o["price"]*o["quantity"] for o in offs)
        ws1.cell(row=row, column=6, value=total).font = Font(bold=True)
    aw(ws1)

    # ── 2. Oylik ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Oylik")
    hdr(ws2, ["Oy", "Bitimlar", "Jami savdo (so'm)"], "2E7D32")
    monthly = await db_all(
        "SELECT SUBSTR(o.created_at,1,7) as month, "
        "COUNT(*) as cnt, SUM(o.price*n.quantity) as total "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "GROUP BY month ORDER BY month DESC LIMIT 24",
        (uid,)
    )
    for i, m in enumerate(monthly, 2):
        ws2.cell(row=i, column=1, value=m["month"])
        ws2.cell(row=i, column=2, value=m["cnt"])
        ws2.cell(row=i, column=3, value=m["total"])
    aw(ws2)

    # ── 3. Mahsulotlar ───────────────────────────────────────────
    ws3 = wb.create_sheet("Mahsulotlar")
    hdr(ws3, ["Mahsulot", "Bitimlar", "Jami (so'm)", "O'rtacha narx"], "1565C0")
    prods = await db_all(
        "SELECT n.product_name, COUNT(*) as cnt, "
        "SUM(o.price*n.quantity) as total, AVG(o.price) as avg_p "
        "FROM offers o JOIN needs n ON o.need_id=n.id "
        "WHERE o.seller_id=? AND o.status=\'accepted\' "
        "GROUP BY n.product_name ORDER BY total DESC",
        (uid,)
    )
    for i, p in enumerate(prods, 2):
        ws3.cell(row=i, column=1, value=p["product_name"])
        ws3.cell(row=i, column=2, value=p["cnt"])
        ws3.cell(row=i, column=3, value=p["total"])
        ws3.cell(row=i, column=4, value=round(p["avg_p"], 0) if p["avg_p"] else 0)
    aw(ws3)

    path = os.path.join(BASE_DIR, f"seller_{uid}_{now.strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(path)
    return path

def calc_ad_price(regions: list, audiences: list) -> tuple:
    """Ball narxini hisoblaydi. (jami, izoh)"""
    multiplier = 2 if len(audiences) >= 2 else 1
    total = 0
    details = []
    for reg in regions:
        base = AD_REGION_PRICES.get(reg, AD_REGION_DEFAULT)
        price = base * multiplier
        total += price
        if multiplier == 2:
            details.append(f"{reg}: {base}×2={price} ball")
        else:
            details.append(f"{reg}: {price} ball")
    return total, details


async def _send_order_to_group(shop, order_id: int, msg_txt: str,
                               buyer_id: int, total: float):
    """Agar do'konda guruh bog'liq bo'lsa, guruhga ham yuboradi."""
    if not shop or not shop.get("group_chat_id"):
        return
    group_id = shop["group_chat_id"]
    claim_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="✋ Men olaman",
            callback_data=f"claim_order_{order_id}"
        )
    ]])
    try:
        await bot.send_message(group_id, msg_txt, reply_markup=claim_kb)
    except Exception as e:
        log.error(f"Guruhga yuborish xato: {e}")

async def _notify_winner(seller_id: int, clinic: dict, items: list):
    """G'olib sotuvchiga to'liq xabar: mahsulotlar + klinika ma'lumoti."""
    cname   = clinic["clinic_name"] or clinic["full_name"] or "Klinika"
    cphone  = clinic["phone"] or "—"
    cregion = clinic["region"] or "—"
    caddr   = clinic["address"] or "—"
    clat    = clinic.get("latitude")
    clon    = clinic.get("longitude")

    # Mahsulotlar ro'yxati
    lines  = []
    total  = 0
    for i, (name, qty, unit, price) in enumerate(items, 1):
        subtotal = price * qty
        total   += subtotal
        lines.append(
            f"{i}. *{name}* — {qty} {unit}\n"
            f"   {price:,.0f} × {qty} = *{subtotal:,.0f} so'm*"
        )
    items_txt = "\n".join(lines)

    # Sotuvchining to'lov usullari
    seller_u = await get_user(seller_id)
    pay_icons = {"p2p":"💳 P2P","cash":"💵 Naqd","bank":"🏦 Hisob raqam"}
    spm_raw = (seller_u.get("payment_methods") or "") if seller_u else ""
    spm_txt = " · ".join(pay_icons[p] for p in spm_raw.split(",") if p in pay_icons)
    spm_line = f"\n💳 To\'lov: {spm_txt}" if spm_txt else ""

    txt = (
        f"🎉 *Taklifingiz qabul qilindi!*\n\n"
        f"📦 *Buyurtma:*\n{items_txt}\n\n"
        f"💰 *Jami: {total:,.0f} so\'m*{spm_line}\n\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🏥 *{cname}*\n"
        f"📞 {cphone}\n"
        f"📍 {cregion}\n"
        f"🏠 {caddr}"
    )

    try:
        await bot.send_message(seller_id, txt)
    except Exception as e:
        log.error(f"Winner notify xato: {e}")
        return

    # Lokatsiya yuboramiz (agar bo'lsa)
    if clat and clon:
        try:
            await bot.send_location(seller_id, latitude=clat, longitude=clon)
        except Exception:
            pass

async def _notify_loser(seller_id: int, product_name: str, win_price: float,
                        my_price: float, unit: str):
    """Yutqazgan sotuvchiga faqat narx statistikasi."""
    diff = my_price - win_price
    pct  = abs(diff) / win_price * 100 if win_price else 0
    try:
        await bot.send_message(
            seller_id,
            f"📊 *{product_name}* bo\'yicha boshqa taklif qabul qilindi.\n\n"
            f"Qabul qilingan narx: *{win_price:,.0f} so\'m/{unit}*\n"
            f"Sizning narxingiz: *{my_price:,.0f} so\'m/{unit}*\n"
            f"Farq: *{diff:+,.0f} so\'m ({pct:.0f}%)*\n\n"
            f"_Xaridor ma\'lumotlari maxfiy._"
        )
    except Exception:
        pass

async def _show_delivery_method(msg_or_obj, state: FSMContext):
    """Yetkazish usulini tanlash."""
    await state.set_state(CheckoutState.delivery)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text="🚗 Sotuvchi o'zi yetkazadi",
            callback_data="co_delivery_self"
        )],
        [InlineKeyboardButton(
            text="📦 BTS Express orqali",
            callback_data="co_delivery_bts"
        )],
    ])
    txt = (
        "🚚 *Yetkazib berish usuli:*\n\n"
        "🚗 *Sotuvchi o'zi* — Narx sotuvchi belgilaydi\n"
        "📦 *BTS Express* — Pochta orqali"
    )
    if hasattr(msg_or_obj, 'message'):
        await msg_or_obj.message.answer(txt, reply_markup=kb)
    else:
        await msg_or_obj.answer(txt, reply_markup=kb)

async def _show_checkout_confirm(msg, state: FSMContext, data: dict, method: str):
    """Buyurtmani tasdiqlash sahifasi."""
    addr      = data.get("delivery_address", "—")
    phone     = data.get("delivery_phone", "—")
    recipient = data.get("delivery_recipient", "—")
    method_txt = "🚗 Sotuvchi o'zi yetkazadi" if method == "seller_self" else "📦 BTS Express"

    txt = (
        "📋 *Buyurtmani tasdiqlash*\n\n"
        f"📍 Manzil: `{addr}`\n"
        f"👤 Qabul qiluvchi: *{recipient}*\n"
        f"📞 Telefon: `{phone}`\n"
        f"🚚 Yetkazish: {method_txt}\n\n"
        "💳 *To'lov usuli:*"
    )

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🚚 Yetkazilganda (COD)", callback_data="co_pay_cod"),
        ],
        [
            InlineKeyboardButton(text="✏️ Manzilni o'zgartirish", callback_data="co_edit_addr"),
            InlineKeyboardButton(text="❌ Bekor", callback_data="co_cancel"),
        ]
    ])
    await state.set_state(CheckoutState.confirm)
    await msg.answer(txt, reply_markup=kb)

async def _save_review(buyer_id: int, state: FSMContext, comment: str):
    d         = await state.get_data()
    order_id  = d.get("order_id")
    seller_id = d.get("seller_id")
    rating    = d.get("rating", 5)
    await state.clear()

    await db_insert(
        "INSERT INTO reviews(order_id,buyer_id,seller_id,rating,comment) VALUES(?,?,?,?,?)",
        (order_id, buyer_id, seller_id, rating, comment)
    )
    # Do'kon reytingini yangilaymiz
    avg = await db_get(
        "SELECT AVG(rating) as avg FROM reviews WHERE seller_id=?", (seller_id,)
    )
    if avg and avg["avg"]:
        await db_run(
            "UPDATE shops SET rating=? WHERE owner_id=?",
            (round(float(avg["avg"]), 1), seller_id)
        )

async def delivery_checker():
    """Har soat ishlaydigan — 48 soat o'tgan buyurtmalarni tekshiradi."""
    while True:
        await asyncio.sleep(3600)  # Har soat
        try:
            # 48 soat o'tgan, hali notify yuborilmagan confirmed buyurtmalar
            cutoff = (datetime.now() - timedelta(hours=48)).isoformat()
            orders = await db_all(
                "SELECT * FROM catalog_orders "
                "WHERE status='confirmed' AND notify_sent=0 "
                "AND confirmed_at <= ?",
                (cutoff,)
            )
            for order in orders:
                try:
                    delivery_kb = InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(
                            text="✅ Ha, yetib keldi!",
                            callback_data=f"co_delivered_{order['id']}_{order['seller_id']}"
                        ),
                        InlineKeyboardButton(
                            text="❌ Yo'q, muammo bor",
                            callback_data=f"co_problem_{order['id']}_{order['seller_id']}"
                        )
                    ]])
                    # Mahsulotlar ro'yxati
                    import json as _pj
                    try:
                        items = _pj.loads(order["products_json"] or "[]")
                        prod_txt = ", ".join([f"{i['name']} x{i['qty']}" for i in items[:3]])
                    except Exception:
                        prod_txt = "mahsulotlar"

                    await bot.send_message(
                        order["buyer_id"],
                        f"📦 *Buyurtmangiz yetib keldimi?*\n\n"
                        f"_{prod_txt}_\n\n"
                        f"Iltimos, tasdiqlang:",
                        reply_markup=delivery_kb
                    )
                    await db_run(
                        "UPDATE catalog_orders SET notify_sent=1 WHERE id=?",
                        (order["id"],)
                    )
                    await asyncio.sleep(0.1)
                except Exception as e:
                    log.error(f"Delivery notify xato order {order['id']}: {e}")
        except Exception as e:
            log.error(f"delivery_checker xato: {e}")

async def _show_product_start(msg: Message, pid: int):
    """Deep link orqali kelgan foydalanuvchiga mahsulotni ko'rsatish."""
    prod = await db_get(
        "SELECT p.*, s.shop_name, s.owner_id as seller_id, u.region "
        "FROM products p "
        "JOIN shops s ON p.shop_id=s.id "
        "JOIN users u ON s.owner_id=u.id "
        "WHERE p.id=? AND p.is_active=1",
        (pid,)
    )
    if not prod:
        await msg.answer("❌ Mahsulot topilmadi yoki o\'chirilgan.")
        return

    stars = ""
    avg = await db_get(
        "SELECT AVG(rating) as a, COUNT(*) as c FROM reviews WHERE seller_id=?",
        (prod["seller_id"],)
    )
    if avg and avg["a"]:
        rating = float(avg["a"])
        stars = "⭐" * round(rating) + f" ({rating:.1f})"

    txt = (
        f"🦷 *{prod['name']}*\n\n"
        f"💰 *{prod['price']:,.0f} so\'m / {prod['unit']}*\n"
        f"🏪 {prod['shop_name']}\n"
        f"📍 {prod['region'] or '—'}\n"
        + (f"⭐ {stars}\n" if stars else "") +
        (f"\n_{prod['description']}_" if prod.get("description") else "")
    )
    kb = ik(
        [ib("⚡ Tezkor buyurtma", f"qorder_{pid}_{prod['seller_id']}")],
        [ib("🔳 QR kod olish", f"get_qr_{pid}"),
         ib("📤 Ulashish", f"share_prod_{pid}")],
    )
    # Rasm bor bo'lsa
    photo = await db_get(
        "SELECT file_id FROM product_photos WHERE product_id=? ORDER BY sort_order LIMIT 1",
        (pid,)
    )
    if photo:
        try:
            await msg.answer_photo(photo["file_id"], caption=txt, reply_markup=kb)
            return
        except Exception: pass
    await msg.answer(txt, reply_markup=kb)

async def _send_quick_order(target_msg, state: FSMContext, qty: float, buyer_id: int):
    d         = await state.get_data()
    pid       = d.get("qo_pid")
    seller_id = d.get("qo_seller")
    name      = d.get("qo_name", "?")
    price     = d.get("qo_price", 0)
    unit      = d.get("qo_unit", "dona")
    total     = price * qty
    await state.clear()

    u      = await get_user(buyer_id)
    uname  = (u["clinic_name"] or u["full_name"] or str(buyer_id)) if u else str(buyer_id)
    uphone = u["phone"] if u else "—"
    uregion= u["region"] if u else "—"
    uaddr  = u["address"] if u else "—"

    # catalog_orders ga yozamiz
    import json as _pj
    items = [{"name": name, "qty": qty, "price": price, "unit": unit, "subtotal": total}]
    order_id = await db_insert(
        "INSERT INTO catalog_orders(buyer_id,seller_id,products_json,total_amount) VALUES(?,?,?,?)",
        (buyer_id, seller_id, _pj.dumps(items, ensure_ascii=False), total)
    )

    msg_txt = (
        f"⚡ *Tezkor buyurtma #{order_id}!*\n\n"
        f"📦 *{name}* — {qty} {unit}\n"
        f"💰 {price:,.0f} × {qty} = *{total:,.0f} so\'m*\n\n"
        f"━━━━━━━━━━━━━━━━━━\n"
        f"🏥 *{uname}*\n"
        f"📞 {uphone}\n"
        f"📍 {uregion}\n"
        f"🏠 {uaddr}"
    )
    confirm_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Qabul qildim",
                             callback_data=f"co_confirm_{order_id}_{buyer_id}"),
        InlineKeyboardButton(text="❌ Mavjud emas",
                             callback_data=f"co_reject_{order_id}_{buyer_id}")
    ]])
    try:
        await bot.send_message(seller_id, msg_txt, reply_markup=confirm_kb)
    except Exception as e:
        log.error(f"Quick order notify xato: {e}")

    # Guruhga ham
    shop_g = await db_get("SELECT * FROM shops WHERE owner_id=?", (seller_id,))
    await _send_order_to_group(
        dict(shop_g) if shop_g else None,
        order_id, msg_txt, buyer_id, total
    )

    await target_msg.answer(
        f"✅ *Buyurtma #{order_id} yuborildi!*\n\n"
        f"📦 {name} — {qty} {unit}\n"
        f"💰 *{total:,.0f} so\'m*\n\n"
        f"Sotuvchi tez orada bog\'lanadi."
    )

async def _generate_article_code() -> str:
    """XD00001 dan boshlanadi, to'lganda avtomatik uzayadi."""
    row = await db_get(
        "SELECT article_code FROM products "
        "WHERE article_code IS NOT NULL "
        "ORDER BY LENGTH(article_code) DESC, article_code DESC LIMIT 1"
    )
    if row and row["article_code"]:
        try:
            last_num = int(row["article_code"][2:])
            next_num = last_num + 1
            # Minimum 5 xona, kerak bo'lsa uzayadi
            digits = max(5, len(str(next_num)))
            return f"XZ{next_num:0{digits}d}"
        except Exception:
            pass
    return "XZ00001"

def _generate_qr_bytes(url: str, label: str = "") -> bytes:
    """QR kod PNG — ichida artikul kodi yozilgan."""
    try:
        import qrcode
        import io
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)

        try:
            from PIL import Image, ImageDraw, ImageFont
            qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
            if label:
                qr_w, qr_h = qr_img.size
                txt_h = 52
                final = Image.new("RGB", (qr_w, qr_h + txt_h), "white")
                final.paste(qr_img, (0, 0))
                draw = ImageDraw.Draw(final)
                # Font qidirish
                font = None
                font_paths = [
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                    "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
                ]
                for fp in font_paths:
                    try:
                        font = ImageFont.truetype(fp, 26)
                        break
                    except Exception:
                        continue
                if font is None:
                    font = ImageFont.load_default()
                try:
                    bbox = draw.textbbox((0, 0), label, font=font)
                    txt_w = bbox[2] - bbox[0]
                except Exception:
                    txt_w = len(label) * 14
                x = max(0, (qr_w - txt_w) // 2)
                y = qr_h + 10
                draw.text((x, y), label, fill="#090979", font=font)
                # XAZDENT yozuvi
                try:
                    bbox2 = draw.textbbox((0, 0), "XAZDENT", font=font)
                    tw2 = bbox2[2] - bbox2[0]
                except Exception:
                    tw2 = 80
                draw.text(((qr_w - tw2)//2, qr_h - 2), "XAZDENT", fill="#444DCF", font=font)
            else:
                final = qr_img
        except Exception as pil_e:
            log.warning(f"PIL xato, oddiy QR: {pil_e}")
            qr_img = qr.make_image(fill_color="black", back_color="white")
            final = qr_img

        buf = io.BytesIO()
        final.save(buf, format="PNG")
        return buf.getvalue()
    except ImportError as e:
        raise Exception(f"qrcode kutubxonasi o'rnatilmagan: {e}")

async def _send_product_qr(user_id: int, product_id: int):
    """Foydalanuvchiga mahsulot QR kodini yuboradi."""
    prod = await db_get("SELECT * FROM products WHERE id=?", (product_id,))
    if not prod:
        return
    code = prod["article_code"] or f"p_{product_id}"
    url  = f"https://t.me/XazdentBot?start=xd_{code}"
    try:
        qr_bytes = _generate_qr_bytes(url, label=code)
        caption  = (
            f"🔳 *{prod['name']}* QR kodi\n\n"
            f"📌 Artikul: *{code}*\n"
            f"🔗 Havola: `{url}`\n\n"
            f"_Instagram postga joylashtiring — "
            f"mijozlar skan qilib buyurtma beradi_"
        )
        await bot.send_photo(
            user_id,
            BufferedInputFile(qr_bytes, filename=f"{code}_qr.png"),
            caption=caption
        )
    except Exception as e:
        log.error(f"QR yuborish xato: {e}")

async def _post_order_to_group(order_id: int, shop: dict,
                               products_txt: str, total: float,
                               buyer_name: str, buyer_region: str):
    """Buyurtmani do'kon guruhiga yuboradi."""
    group_id = shop.get("group_chat_id")
    if not group_id:
        return
    shop_name = shop.get("shop_name", "Do'kon")
    msg_txt = (
        f"🛒 *Yangi buyurtma #{order_id}!*\n\n"
        f"📦 {products_txt}\n\n"
        f"💰 *Jami: {total:,.0f} so\'m*\n\n"
        f"🏥 *{buyer_name}*\n"
        f"📍 {buyer_region}\n\n"
        f"_Kim qabul qiladi?_"
    )
    claim_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="✋ Men olaman",
            callback_data=f"claim_{order_id}"
        )
    ]])
    try:
        sent = await bot.send_message(group_id, msg_txt, reply_markup=claim_kb)
        await db_run(
            "UPDATE catalog_orders SET group_message_id=? WHERE id=?",
            (sent.message_id, order_id)
        )
    except Exception as e:
        log.error(f"Guruhga post xato (group={group_id}): {e}")

async def _get_usd_rate_from_cbu():
    """CBU dan USD kursini oladi."""
    try:
        import aiohttp as _ah
        async with _ah.ClientSession() as session:
            async with session.get(
                "https://cbu.uz/uz/arkhiv-kursov-valyut/json/USD/",
                timeout=_ah.ClientTimeout(total=10),
                ssl=False
            ) as resp:
                if resp.status == 200:
                    data = await resp.json(content_type=None)
                    if data and len(data) > 0:
                        rate = float(data[0].get("Rate", 0))
                        if rate > 1000:  # Mantiqiy tekshiruv
                            return rate
    except Exception as e:
        log.error(f"CBU kurs olish xato: {e}")
    return None

async def usd_rate_checker():
    """Har kuni 09:00 da USD kursini yangilaydi."""
    import asyncio
    while True:
        try:
            from datetime import datetime as _dt
            now = _dt.now()
            # Keyingi 09:00 gacha kutamiz
            next_run = now.replace(hour=9, minute=0, second=0, microsecond=0)
            if now >= next_run:
                from datetime import timedelta as _td
                next_run += _td(days=1)
            wait_secs = (next_run - now).total_seconds()
            log.info(f"💱 USD kurs yangilanishi: {next_run.strftime('%d.%m %H:%M')} da ({int(wait_secs/3600)}h)")
            await asyncio.sleep(wait_secs)

            # CBU dan kurs olish
            rate = await _get_usd_rate_from_cbu()
            if rate:
                await db_run(
                    "UPDATE settings SET value=? WHERE key='usd_rate'",
                    (str(rate),))
                await db_run(
                    "UPDATE settings SET value=? WHERE key='usd_rate_updated'",
                    (_dt.now().strftime("%d.%m.%Y %H:%M"),))
                log.info(f"✅ USD kurs yangilandi: {rate:.2f} so'm")
            else:
                log.warning("⚠️ CBU dan kurs olinmadi — eski qiymat saqlanadi")
        except Exception as e:
            log.error(f"usd_rate_checker xato: {e}")
            await asyncio.sleep(3600)  # Xato bo'lsa 1 soatdan keyin qayta

async def expire_checker():
    """Har 15 daqiqada muddati o'tgan ehtiyojlarni yopadi."""
    while True:
        try:
            now = datetime.now().isoformat()
            expired = await db_all(
                "SELECT * FROM needs WHERE status='active' AND expires_at IS NOT NULL AND expires_at < ?",
                (now,),
            )
            for n in expired:
                await db_run("UPDATE needs SET status='done' WHERE id=?", (n["id"],))
                # Kanal postini o'chirishga harakat
                if n.get("channel_message_id"):
                    try:
                        await bot.delete_message(CHANNEL_ID, n["channel_message_id"])
                    except Exception:
                        pass
            if expired:
                log.info(f"⏰ {len(expired)} ta ehtiyoj muddati tugadi")
        except Exception as e:
            log.error(f"Expire checker xato: {e}")
        await asyncio.sleep(900)  # 15 daqiqa



# ── YORDAM XIZMATI (support) — bot va Mini App uchun umumiy yetkazish ─────────
async def _support_group_id():
    """Support guruh ID: avval DB sozlamasi, keyin SUPPORT_GROUP_ID env."""
    gid = await get_setting("support_group_id")
    if gid:
        return gid
    env = os.getenv("SUPPORT_GROUP_ID", "").strip()
    return env or None

async def _deliver_support(uid, text, username=None, full_name=None):
    """Support xabarini saqlaydi va support guruhi (yoki adminlar)ga yuboradi.
    uid orqali ishlaydi — bot ham, Mini App API ham chaqira oladi."""
    u     = await get_user(uid)
    name  = ((u["clinic_name"] or u["full_name"]) if u else None) or full_name or str(uid)
    phone = (u["phone"] if u and u["phone"] else "—")
    role  = (u["role"] if u and u["role"] else "—")
    uname = ("@" + username) if username else "—"
    mid = await db_insert("INSERT INTO support_messages(user_id, message) VALUES(?,?)", (uid, text))
    out = (
        f"💬 *Yangi yordam xabari #{mid}*\n\n"
        f"👤 {name}  ·  {role}\n"
        f"🆔 `{uid}`   {uname}\n"
        f"📱 {phone}\n\n"
        f"📝 {text}"
    )
    kb  = ik([ib("💬 Javob berish", f"reply_{mid}_{uid}")])
    gid = await _support_group_id()
    if gid:
        try:
            await bot.send_message(int(gid), out, reply_markup=kb)
            return mid
        except Exception as e:
            log.warning(f"support guruhga yuborilmadi ({gid}): {e}")
    for aid in ADMIN_IDS:
        try:
            await bot.send_message(aid, out, reply_markup=kb)
        except Exception:
            pass
    return mid
